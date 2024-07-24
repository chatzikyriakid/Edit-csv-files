import pandas as pd
import openpyxl as op
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles import Border, Side
from datetime import datetime

def get_original_path():
    path = input('Give the location of the CSV file: ')
    file_name = input('Give the file name: ')
    path = path.replace("\\", "\\\\")
    return path + '\\\\' + file_name

def get_output_path():
    path = input('Give the location of the output file: ')
    file_name = input('Give the output file name: ')
    path = path.replace("\\", "\\\\")
    return path + '\\\\' + file_name

# Function that reads the csv file, removes the uneccessary columns and saves it to an excel file
def read_and_save(original_file,output_file):
    # Reading the CSV file
    df_new = pd.read_csv(original_file, encoding='utf8')

    # filter file based on current date on column Record Date
    current_time = str(datetime.now().strftime("%d/%m/%Y"))
    df_new = df_new[df_new['Record Date'] == current_time + ' 00:00']

    # Delete unnecessary columns
    df_new = df_new.drop(columns=['Batch','Revision','Purchase Order','Quantity Received','Inspector','Action','Inspection Method','Record Date','Entry Date','Id'])

    # Save to xlsx file
    df_new.to_excel(output_file, index=False) 
    print('File saved to {}'.format(output_file))


# Function that formats the excel file (color and borders)
def format_file(path):

    # change cell color
    wb = op.load_workbook(path)
    ws = wb['Sheet1']  

    cell_ids = ['A1', 'B1', 'C1', 'D1']
    for i in range(len(cell_ids)):
        ws[cell_ids[i]].fill = PatternFill(patternType='solid',
                            fgColor=Color(rgb='B2B2B2'))
        
    # add borders
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    wb.save(path)
    print('File saved to {}'.format(path))


#-------------------------------------------------------------------------------
original_path = get_original_path()
output_path = get_output_path()

read_and_save(original_path,output_path)
format_file(output_path)

print("*****************************************************")
input('Press enter to exit')

